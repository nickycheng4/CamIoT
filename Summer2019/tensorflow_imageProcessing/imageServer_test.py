from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import io
import socket
import struct
from PIL import Image
import numpy as np
import cv2

import openpyxl


# ------------------------------------------ #
#               Image Classification         #
# ------------------------------------------ #
import argparse
import sys,os
import time

import numpy as np
import tensorflow as tf

from supportive_functions import load_graph,read_tensor_from_image_file,load_labels


# dictionary for image display
imDic =  {
  "coffee maker": "img_display/cp.jpg",
  "light": "img_display/light.jpg",
  "printer": "img_display/prt.jpg",
  "tv": "img_display/TV.jpg",
  "box": "img_display/TV.jpg",
  "monitor": "img_display/monitor.jpg"
}

distDic = {
	"Close" : "img_display/close.jpg",
  	"Far" : "img_display/far.jpg"
}

now = time.strftime("%x")
now = now.replace('/','-')
now = "Data_Collection/" + now + ".xlsx"

f = open("Data_Collection/test_sample.txt", "r")
sample_file = ("Data_Collection/stat_temp.xlsx") 

wb = openpyxl.load_workbook(filename=sample_file, read_only=False, keep_vba=True)
stat_sheet = wb.worksheets[0]


# ------------------------------------------ #
#           Server Client Interface          #
# ------------------------------------------ #
# Start a socket listening for connections on 0.0.0.0:8000 (0.0.0.0 means
# all interfaces)

connect_start = time.time()
server_socket = socket.socket()
server_socket.bind(('172.31.97.133', 8000))
server_socket.listen(0)

# Accept a single connection and make a file-like object out of it
connection = server_socket.accept()[0].makefile('rb')

connect_finish = time.time()
print ("\n")
print ("Total connection time: %.3f s.\n" %(connect_finish - connect_start))

row_accu = 2
col_num = 4

test_round = 1
numerator = 0.0
demoninator = 0.0

total_result = []
try:
	for item_obj in f:
		#print(item_obj)
		item_obj = item_obj.split()
		#print(item_obj)
		if len(item_obj) > 2:
			item_full = " ".join(item_obj[:len(item_obj)-1])
			item_obj[0] = item_full
		print("Test Round %d: Object %s with distance %s, Please take a picuture.\n" %(test_round,item_obj[0],item_obj[-1]))
		test_round += 1
		# Read the length of the image as a 32-bit unsigned int. If the
		# length is zero, quit the loop
		image_len = struct.unpack('<L', connection.read(struct.calcsize('<L')))[0]
		if not image_len:
			break
		# Construct a stream to hold the image data and read the image
		# data from the connection
		image_stream = io.BytesIO()
		image_stream.write(connection.read(image_len))
		# Rewind the stream, open it as an image with PIL and do some
		# processing on it
		image_stream.seek(0)    # seek(0) - the start of the file
		
		file_bytes = np.asarray(bytearray(image_stream.read()), dtype=np.uint8)
		img = cv2.imdecode(file_bytes, cv2.IMREAD_UNCHANGED)
		#cv2.imwrite('imageIn1.jpg',img)

		set_size = 224
		display_size = 400

		cv2.waitKey(1)
		img2 = img[135:945,135:945]
		img3 = img[270:810,270:810]

		img1_resized = cv2.resize(img , (set_size,set_size))
		img2_resized = cv2.resize(img2, (set_size,set_size))
		img3_resized = cv2.resize(img3, (set_size,set_size))

		cv2.imwrite('imageIn1.jpg',img1_resized)
		cv2.imwrite('imageIn2.jpg',img2_resized)
		cv2.imwrite('imageIn3.jpg',img3_resized)
		cv2.waitKey(1)

		img_list = ["imageIn1.jpg","imageIn2.jpg","imageIn3.jpg"]

		# image classification
		if __name__ == "__main__":
			distance_calc_start = time.time()

			file_name = img_list[0]
			model_file = "tf_files/tf_files_2/retrained_graph.pb"
			label_file = "tf_files/tf_files_2/retrained_labels.txt"
			input_height = set_size
			input_width = set_size
			input_mean = 128
			input_std = 128
			input_layer = "input"
			output_layer = "final_result"

			graph = load_graph(model_file)
			t = read_tensor_from_image_file(file_name,
											input_height=input_height,
											input_width=input_width,
											input_mean=input_mean,
											input_std=input_std)
			input_name = "import/" + input_layer
			output_name = "import/" + output_layer
			input_operation = graph.get_operation_by_name(input_name);
			output_operation = graph.get_operation_by_name(output_name);

			with tf.Session(graph=graph) as sess:
			  start = time.time()
			  results = sess.run(output_operation.outputs[0],
								{input_operation.outputs[0]: t})
			  end=time.time()
			results = np.squeeze(results)

			top_k = results.argsort()[-2:][::-1]
			labels_dist = load_labels(label_file)

			if labels_dist[top_k[0]].lower() == "close":
				img_weight = [4.0,2.0,0.0]
			elif labels_dist[top_k[0]].lower() == "far":
				img_weight = [0.0,2.0,4.0]
			'''
			print(labels_dist)
			print(top_k)
			print(results)
			'''
			cv2.waitKey(2)
			imgdist = cv2.imread(distDic[labels_dist[top_k[0]].capitalize()],1)
			imgdist = cv2.resize(imgdist, (200,200))

			cv2.waitKey(2)

			cv2.imshow("Predicted Distance: %s" %(labels_dist[top_k[0]]),imgdist)
			cv2.moveWindow("Predicted Distance: %s" %(labels_dist[top_k[0]]),50, 550)

			distance_calc_end = time.time()

			#print ("Distance estimation time: %.2f milliseconds." %((distance_calc_end - distance_calc_start)*1000))
			#print ("Predicted distance: %s with confidence %.2f." %(labels_dist[top_k[0]],results[top_k[0]]))


			total_scoring = {}
			for item in imDic.keys():
				total_scoring[item] = 0.0

			object_recog_start = time.time()
			for image in img_list:
				file_name = image
				model_file = "tf_files/tf_files_1/retrained_graph.pb"
				label_file = "tf_files/tf_files_1/retrained_labels.txt"
				input_height = set_size
				input_width = set_size
				input_mean = 128
				input_std = 128
				input_layer = "input"
				output_layer = "final_result"

				graph = load_graph(model_file)
				t = read_tensor_from_image_file(file_name,
												input_height=input_height,
												input_width=input_width,
												input_mean=input_mean,
												input_std=input_std)
				input_name = "import/" + input_layer
				output_name = "import/" + output_layer
				input_operation = graph.get_operation_by_name(input_name);
				output_operation = graph.get_operation_by_name(output_name);

				with tf.Session(graph=graph) as sess:
					start = time.time()
					results = sess.run(output_operation.outputs[0],
									{input_operation.outputs[0]: t})
					end=time.time()
				results = np.squeeze(results)

				top_k = results.argsort()[-3:][::-1]
				labels = load_labels(label_file)

				cur_weight = int(image[-5])-1

				for item in labels:
					total_scoring[item] += img_weight[cur_weight]*results[labels.index(item)]

				if labels[top_k[0]] not in imDic.keys():
					print ("Can not display recognized object as the image can not be found.\n")
					break

				imgDis = cv2.imread(imDic[labels[top_k[0]]],1)

				img_show = cv2.imread('imageIn1.jpg',1)
				imgDis = cv2.resize(imgDis, (display_size,display_size))
				new_img = cv2.resize(img_show,(display_size,display_size))

				cv2.waitKey(2)
				if image == "imageIn1.jpg":
					cv2.imshow('Actual Object',new_img)
					cv2.moveWindow('Actual Object',660,50)

				cv2.waitKey(2)
				
				#print(50+(int(image[-5])-1)*60)
				
				cv2.imshow("Predicted Result for %s" %(image),imgDis)
				cv2.moveWindow("Predicted Result for %s" %(image),50+(int(image[-5])-1)*50 ,50+(int(image[-5])-1)*50)

			factor=1.0/sum(total_scoring.itervalues())
			
			for name in total_scoring:
				total_scoring[name] = total_scoring[name]*factor

			final_scoring = sorted(total_scoring.items(), key=lambda x: x[1], reverse = True)

			object_recog_end = time.time()
			#print ("Object recognition time: %.2f milliseconds." %((object_recog_end - object_recog_start)*1000))

			#print ("Image Weight Matrix: %s.\n" %(img_weight))

			#print(final_scoring)
			template = "{}: (score={:0.3f})"

			cur_result = final_scoring[0][0]
			total_result.append(cur_result)

			for item in final_scoring:
			  print (template.format(item[0].title(),item[1]))

			cv2.waitKey(500) # show inmage for 5 secs
			cv2.destroyAllWindows()
			cv2.waitKey(2)

			print('\n')

			#print (final_scoring)
			#print (item_obj)
			if final_scoring[0][0].lower() == item_obj[0].lower():
				print("CORRECT!!!")
				numerator += 1
			else:
				print("Opps, Try Again...")
				
			demoninator+=1

			print("Current test result: %d out of %d correct.\nPrecision rate: %.2f%% ." %(numerator,demoninator,numerator*100/demoninator))

	print ("Test case completed! See Excel File for statistical document!")
	if total_result:
		for item in total_result:

			if item.lower() == 'tv':
				#print ("1",item.upper())
				stat_sheet.cell(row = row_accu, column = col_num).value = item.upper()
			else:
				#print ("2",item.title())
				stat_sheet.cell(row = row_accu, column = col_num).value = item.title()
			row_accu += 1

		wb.save(now)
	
	

finally:
	connection.close()
	server_socket.close()

f.close()
