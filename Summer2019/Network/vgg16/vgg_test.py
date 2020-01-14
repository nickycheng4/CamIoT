from keras.preprocessing import image
from keras.applications.vgg16 import VGG16
from keras.applications.vgg16 import preprocess_input
from keras.models import Model
import numpy as np
import os
from joblib import load
from openpyxl import Workbook,load_workbook
import time
import cv2
import sys
sys.path.insert(0, 'Supportive_Funcs')
from confusion_matrix import draw_matrix

sys.path.insert(0, 'Finger_Detection')
from crop import generate_crop

# model = VGG16(weights='imagenet', include_top=False)
# model.summary()

now = time.strftime("%x")
now = now.replace('/','-')
now = "Data_Collection/" + now + ".xlsx"

wb=Workbook()
filepath=now
wb.save(filepath)
wb=load_workbook(filepath)
sheet=wb.active

tru_col = 1
pre_col = 2
row_tru = 1
row_pre = 1

total_img_process = 0
crop_ratio_num = float(4)
assigned_dim = 224

red_bias = -40
blue_bias = -20
green_bias = 0

layer = 'fc2'
base_model = VGG16(weights='imagenet')
model = Model(inputs=base_model.input, outputs=base_model.get_layer(layer).output)
clf = load('Trained_SVM/Nic/Raw_6MLP.joblib') 
owd = os.getcwd()

os.chdir('test_datasets')

rootdir = 'Fun'
labels = []
for subdir, dirs, files in os.walk(rootdir):
	#print(subdir,dirs,files)
	test_set = []
	#print(subdir.partition('/'))
	if subdir.partition('/')[-1] =='':
		continue
	correct_tag = subdir.partition('/')[-1]
	labels.append(correct_tag)
	#print(correct_tag)
	print('Now in the folder',subdir.partition('/')[-1])

	files_size = 0
	counter = 0
	for i, file in enumerate(files):
		file_path = os.path.join(subdir, file)
		if 'jpg' in file_path or 'JPG' in file_path or 'Screen Shot' in file_path:
			total_img_process +=1
			files_size+=1
			print('\nImage: %s'%(file_path))
			img_crop,img_bk = generate_crop(file_path,220)
			'''
			height,width,channel = img_crop.shape
			for line in range(height):
				for pixel in range(width):
					#print((line,pixel),(img_1[line,pixel][1],img_2[line,pixel][1],max([img_1[line,pixel][1],img_2[line,pixel][1]])-min([img_1[line,pixel][1],img_2[line,pixel][1]])))
					if img_crop[line,pixel][2] > abs(red_bias):
						img_crop[line,pixel][2] += red_bias
					else:
						img_crop[line,pixel][2]  = 0

					if img_crop[line,pixel][0] > abs(blue_bias):
						img_crop[line,pixel][0] += blue_bias
					else:
						img_crop[line,pixel][0] = 0 
					
					if img_crop[line,pixel][1] > abs(green_bias)
						img_crop[line,pixel][1] += green_bias
					else:
						img_crop[line,pixel][1] = 0
					
			#cv2.imwrite('0.jpg',img_crop)
			'''
			bkimg_dir = rootdir+'/bk'+file
			cpimg_dir = rootdir+'/cp'+file
			
			#print((bkimg_dir,cpimg_dir))
			cv2.imwrite(bkimg_dir,img_bk)
			cv2.imwrite(cpimg_dir,img_crop)
			img = image.load_img(cpimg_dir, target_size=(224, 224))
			#img = generate_crop(file_path,240)

			img_data = image.img_to_array(img)
			img_data = np.expand_dims(img_data, axis=0)
			img_data = preprocess_input(img_data)

			vgg16_feature = model.predict(img_data)
			test_set.append(np.ndarray.tolist(vgg16_feature[0]))
			
			#print ('\r>> preprocessing %3.2f' % ((i/files_size)*100),'%',end='.')
	if test_set:
		predict_target = clf.predict(test_set)
		predict_prob = clf.predict_proba(test_set)
		#print(correct_tag)
		print('predict results.')
		print(predict_target)
		predict_target_ls = np.ndarray.tolist(predict_target)
		accuracy = (predict_target_ls.count(correct_tag) / files_size ) * 100
		print('accuracy',accuracy,'%')
		print('probability.')
		print('classes',clf.classes_)
		print(predict_prob)
		for item in predict_target:
			sheet.cell(row=row_pre, column=pre_col).value = str(item)
			sheet.cell(row=row_tru, column=tru_col).value = str(correct_tag)
			row_tru += 1
			row_pre += 1
	
os.chdir(owd)

wb.save(filepath)

draw_matrix(total_img_process,filepath,labels)