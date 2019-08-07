import numpy as np
import matplotlib.pyplot as plt
import openpyxl

from sklearn import svm, datasets
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix
from sklearn.utils.multiclass import unique_labels

from sklearn import metrics
import itertools
import os
import shutil


def plot_confusion_matrix(y_true, y_pred, classes,
						  normalize=False,
						  title=None,
						  cmap=plt.cm.Blues):
	"""
	This function prints and plots the confusion matrix.
	Normalization can be applied by setting `normalize=True`.
	"""
	'''
	if not title:
		if normalize:
			title = 'Normalized confusion matrix'
		else:
			title = 'Confusion matrix, without normalization'
	'''
	# Compute confusion matrix
	cm = confusion_matrix(y_true, y_pred)
	# Only use the labels that appear in the data
	#classes = classes[unique_labels(y_true, y_pred)]
	if normalize:
		cm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
		

	#print(cm)
	fig, ax = plt.subplots()
	im = ax.imshow(cm, interpolation='nearest', cmap=cmap)
	ax.figure.colorbar(im, ax=ax)
	# We want to show all ticks...
	ax.set(xticks=np.arange(cm.shape[1]),
		   yticks=np.arange(cm.shape[0]),
		   # ... and label them with the respective list entries
		   xticklabels=classes, yticklabels=classes,
		   title=title,
		   ylabel='True label',
		   xlabel='Predicted label')

	# Rotate the tick labels and set their alignment.
	plt.setp(ax.get_xticklabels(), rotation=45, ha="right",
			 rotation_mode="anchor")

	# Loop over data dimensions and create text annotations.
	fmt = '.2f' if normalize else 'd'
	thresh = cm.max() / 2.
	for i in range(cm.shape[0]):
		for j in range(cm.shape[1]):
			ax.text(j, i, format(cm[i, j], fmt),
					ha="center", va="center",
					color="white" if cm[i, j] > thresh else "black")
	fig.tight_layout()
	return ax
def draw_matrix(total_num,file_direc,labels,save_name):
  #print(total_num)
  total_line = total_num
  gt=[]
  pd = []
  wb = openpyxl.load_workbook(filename= file_direc, read_only=False, keep_vba=True)
  sheet = wb.active

  cells = sheet['A1': 'A%s'%(str(total_line))]
  for c1 in cells:
	if c1 !='':
	#print(c1[0].value)
	  gt.append(str(c1[0].value.lower()))
  cells_2 = sheet['B1': 'B%s'%(str(total_line))]

  for c2 in cells_2:
	pd.append(str(c2[0].value.lower()))

  #print (gt)
  #print (pd)

  #class_names = ['TV','Printer','Monitor', 'Lamp', 'Coffee Maker']
  class_names = labels
  class_names.sort()

  np.set_printoptions(precision=2)
  
  # Plot non-normalized confusion matrix
  '''
  plot_confusion_matrix(gt, pd, classes=class_names,
						title='Confusion matrix, without normalization')
  '''
  # Plot normalized confusion matrix
  plot_confusion_matrix(gt, pd, classes=class_names, normalize=True,
						title='Normalized confusion matrix')
  #plt.axis('off')
  fig = plt.gcf()
  fig.set_size_inches(len(labels)*0.8, len(labels)*0.8)
  fig.savefig(save_name,bbox_inches='tight')
  plt.close()
  #plt.savefig(save_name,bbox_inches='tight',pad_inches = 0)

  #plt.show()


if __name__ == "__main__":
	names = ['Amirali','Robin','Xue','Liu','Nic','Kazem','Seyed','Mahdi','Zijie','Evirgen']
	image_types = ['Pi','Phone']
	train_types = ['Aug','Raw']
	test_types = ['Fin','Nofin']
	result_dir = 'Result'
	for person in names:
		if person in os.listdir(result_dir):
			shutil.rmtree(result_dir+'/'+person)
		os.mkdir(result_dir+'/'+person)
		for image_type in image_types:
			os.mkdir(result_dir+'/'+person+'/'+image_type)
			for train_type in train_types:
				for test_type in test_types:
					save_name = 'Result/' + person +'/' + image_type +'/'+train_type+' '+test_type +'.jpg'
					excel_dir = 'Data_Collection/%s%s%s%s.xlsx'%(person.lower(),image_type.lower(),train_type.lower(),test_type.lower())


					objects = []
					filename = excel_dir
					wb = openpyxl.load_workbook(filename= filename, read_only=False, keep_vba=True)
					sheet = wb.active
					tru_col = 1
					pre_col = 2
					row_tru = 1
					row_pre = 1


					while sheet.cell(row=row_tru, column=tru_col).value:
					#print(sheet.cell(row=row_tru, column=tru_col).value)
						if sheet.cell(row=row_tru, column=tru_col).value.lower() not in objects:
						  objects.append(str(sheet.cell(row=row_tru, column=tru_col).value.lower()))
						row_tru += 1

					objects_s = sorted(objects, key=lambda s: s.lower())

					#print(objects_s)
					#print(row_tru)

					draw_matrix(row_pre-1,filename,objects_s,save_name)